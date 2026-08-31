import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import os

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeout
from playwright.sync_api import sync_playwright


# -----------------------------------------------------------------------------
# CONFIGURACAO
# -----------------------------------------------------------------------------

RIOT_ID = "elsewhere#999t"
PLATAFORMA = "pc"
PLAYLIST = "competitive"
SEASON_ID = "8102cd81-43a0-d0d7-bd59-47b8fe9bed1b"

MODO_HEADLESS = os.getenv("CI", "").lower() == "true"

PASTA_SAIDA = Path("dados_tracker")


def montar_urls():
    perfil = quote(RIOT_ID, safe="")
    base = f"https://tracker.gg/valorant/profile/riot/{perfil}"
    parametros = (
        f"platform={PLATAFORMA}"
        f"&playlist={PLAYLIST}"
        f"&season={SEASON_ID}"
    )

    return {
        "Agente": f"{base}/agents?{parametros}",
        "Mapa": f"{base}/maps?{parametros}",
    }


def limpar_texto(texto):
    return re.sub(r"\s+", " ", str(texto or "")).strip()


def converter_valor(valor):
    """Converte numeros do Tracker quando for seguro fazer isso."""
    if valor is None:
        return None

    original = limpar_texto(valor)

    if not original or original in {"-", "--", "N/A"}:
        return None

    sem_percentual = original.removesuffix("%").strip()

    # Padrao de milhar ingles: 1,144 -> 1144
    if re.fullmatch(r"-?\d{1,3}(,\d{3})+", sem_percentual):
        return int(sem_percentual.replace(",", ""))

    # Inteiro ou decimal: 42, 0.87, -13
    if re.fullmatch(r"-?\d+(\.\d+)?", sem_percentual):
        numero = float(sem_percentual)
        return int(numero) if numero.is_integer() else numero

    # Tempo e placares W/L permanecem como texto.
    return original


def classes(elemento):
    return elemento.get("class") or []


def filhos_com_classe(elemento, nome_classe):
    if not elemento:
        return []

    return [
        filho
        for filho in elemento.find_all(recursive=False)
        if nome_classe in classes(filho)
    ]


def aceitar_cookies(page):
    textos = [
        "Accept All",
        "Accept all",
        "Accept",
        "I Agree",
        "Allow all",
        "Aceitar tudo",
        "Aceitar",
        "Concordo",
    ]

    for texto in textos:
        try:
            botao = page.get_by_role("button", name=texto, exact=True).first
            if botao.is_visible(timeout=700):
                botao.click(timeout=2_000)
                page.wait_for_timeout(500)
                print("Aviso de cookies fechado.")
                return
        except Exception:
            pass


def pagina_bloqueada(page):
    try:
        texto = limpar_texto(page.locator("body").inner_text()).lower()
    except Exception:
        return False

    mensagens = (
        "verify you are human",
        "verifying you are human",
        "checking your browser",
        "just a moment",
        "attention required",
        "access denied",
        "confirme que voce e humano",
        "verifique se voce e humano",
    )
    return any(mensagem in texto for mensagem in mensagens)


def aguardar_conteudo(page):
    try:
        page.wait_for_load_state("domcontentloaded", timeout=60_000)
    except PlaywrightTimeout:
        print("Aviso: o carregamento inicial excedeu 60 segundos.")

    aceitar_cookies(page)

    try:
        page.wait_for_selector(
            ".st-content .st-content__item",
            state="attached",
            timeout=30_000,
        )
    except PlaywrightTimeout:
        print("Aviso: a tabela ainda nao apareceu; continuarei a verificacao.")

    # Forca o carregamento de linhas e imagens que usam lazy loading.
    for _ in range(6):
        page.mouse.wheel(0, 1_200)
        page.wait_for_timeout(450)

    page.mouse.wheel(0, -10_000)
    page.wait_for_timeout(1_500)


def valor_da_celula(item):
    elemento = item.select_one(".info > .value")

    if not elemento:
        elemento = item.select_one(".value")

    if elemento:
        return limpar_texto(elemento.get_text(" ", strip=True))

    imagem = item.select_one("img[alt]")
    if imagem:
        return limpar_texto(imagem.get("alt"))

    return limpar_texto(item.get_text(" ", strip=True))


def extrair_top_agents(item):
    """Extrai nomes e percentuais da coluna Top Agents dos mapas."""
    agentes = []

    for imagem in item.select("img[alt]"):
        nome = limpar_texto(imagem.get("alt"))
        bloco = imagem.find_parent(
            "div",
            class_=lambda valor: valor and "flex-col" in valor,
        )
        percentual = ""

        if bloco:
            textos = bloco.select("span")
            if textos:
                percentual = limpar_texto(textos[-1].get_text(" ", strip=True))

        if nome:
            agentes.append(f"{nome} ({percentual})" if percentual else nome)

    return "; ".join(agentes)


def extrair_tabelas_tracker(html, tipo):
    """Extrai as tabelas feitas com divs usadas atualmente pelo Tracker.gg."""
    soup = BeautifulSoup(html, "html.parser")
    resultados = []
    primeiras_colunas_esperadas = (
        {"agent"} if tipo == "Agente" else {"map", "map name"}
    )

    for tabela in soup.select(".st"):
        cabecalho = tabela.select_one(":scope > .st-header")
        if not cabecalho:
            cabecalho = tabela.select_one(".st-header")

        itens_cabecalho = filhos_com_classe(cabecalho, "st__item")
        colunas = []

        for item in itens_cabecalho:
            label = item.select_one(".label")
            nome = limpar_texto(label.get_text(" ", strip=True)) if label else ""
            if nome:
                colunas.append(nome)

        if not colunas or colunas[0].lower() not in primeiras_colunas_esperadas:
            continue

        conteudo = tabela.select_one(":scope > .st-content")
        if not conteudo:
            conteudo = tabela.select_one(".st-content")

        if not conteudo:
            continue

        linhas = conteudo.select(".st-content__category > .st-content__item")
        if not linhas:
            linhas = conteudo.select(":scope > .st-content__item")

        for linha in linhas:
            itens = [
                filho
                for filho in linha.find_all(recursive=False)
                if "st-content__item-value" in classes(filho)
            ]

            if not itens:
                continue

            registro = {"Tipo": tipo}

            for indice, item in enumerate(itens):
                if indice >= len(colunas):
                    break

                coluna = colunas[indice]
                if tipo == "Mapa" and coluna == "Top Agents":
                    valor = extrair_top_agents(item)
                else:
                    valor = valor_da_celula(item)
                registro[coluna] = converter_valor(valor)

                if indice == 0 and tipo == "Agente":
                    funcao = item.select_one(".info > .label")
                    if funcao:
                        registro["Funcao"] = limpar_texto(
                            funcao.get_text(" ", strip=True)
                        )

            nome = (
                registro.get("Agent")
                or registro.get("Map")
                or registro.get("Map Name")
            )
            if nome:
                resultados.append(registro)

        if resultados:
            break

    # Remove eventuais repeticoes sem alterar a ordem original.
    unicos = []
    vistos = set()

    for registro in resultados:
        chave = json.dumps(registro, sort_keys=True, ensure_ascii=False, default=str)
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(registro)

    return unicos


def salvar_diagnostico(page, tipo):
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
    identificador = tipo.lower()
    caminho_html = PASTA_SAIDA / f"debug_{identificador}.html"
    caminho_png = PASTA_SAIDA / f"debug_{identificador}.png"

    html = page.content()
    caminho_html.write_text(html, encoding="utf-8")

    try:
        page.screenshot(path=str(caminho_png), full_page=True)
    except Exception as erro:
        print(f"Aviso: nao foi possivel salvar a captura: {erro}")

    return html, caminho_html, caminho_png


def extrair_pagina(page, url, tipo):
    print("\n" + "=" * 72)
    print(f"Abrindo pagina de {tipo}:")
    print(url)

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=90_000)
    except PlaywrightTimeout:
        print("Aviso: a navegacao excedeu 90 segundos.")

    aguardar_conteudo(page)

    if pagina_bloqueada(page):
        print("\nO Tracker esta solicitando uma verificacao humana.")
        print("Conclua a verificacao na janela do navegador.")
        input("Depois, pressione ENTER neste terminal para continuar... ")
        aguardar_conteudo(page)

    html, caminho_html, caminho_png = salvar_diagnostico(page, tipo)
    registros = extrair_tabelas_tracker(html, tipo)

    if registros:
        print(f"Registros encontrados para {tipo}: {len(registros)}")
    else:
        print(f"Nenhum registro de {tipo} foi reconhecido.")
        print(f"HTML de diagnostico: {caminho_html.resolve()}")
        print(f"Imagem de diagnostico: {caminho_png.resolve()}")

    return registros


def organizar_colunas(df):
    if df.empty:
        return df

    primeiras = [
        coluna
        for coluna in ("Tipo", "Agent", "Map", "Map Name", "Funcao", "Top Agents")
        if coluna in df.columns
    ]
    restantes = [coluna for coluna in df.columns if coluna not in primeiras]
    return df[primeiras + restantes]


def formatar_planilha(planilha, dataframe):
    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions
    planilha.sheet_view.showGridLines = False

    preenchimento = PatternFill("solid", fgColor="17212B")

    for celula in planilha[1]:
        celula.fill = preenchimento
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center")

    for coluna in planilha.columns:
        maior = max(len(str(celula.value or "")) for celula in coluna)
        letra = coluna[0].column_letter
        planilha.column_dimensions[letra].width = min(max(maior + 2, 12), 28)

    # Escala visual apenas nas colunas em que valores maiores costumam ser melhores.
    colunas_semaforo = {
        "Win %",
        "K/D",
        "ADR",
        "ACS",
        "DDΔ",
        "HS%",
        "KAST",
        "Attack Win %",
        "Attack K/D",
        "Defense Win %",
        "Defense K/D",
    }

    for indice, nome in enumerate(dataframe.columns, start=1):
        if nome not in colunas_semaforo or len(dataframe) < 2:
            continue

        letra = planilha.cell(row=1, column=indice).column_letter
        intervalo = f"{letra}2:{letra}{len(dataframe) + 1}"
        planilha.conditional_formatting.add(
            intervalo,
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )


def salvar_resultados(agentes, mapas, urls):
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)

    df_agentes = organizar_colunas(pd.DataFrame(agentes))
    df_mapas = organizar_colunas(pd.DataFrame(mapas))

    caminho_agentes = PASTA_SAIDA / "tracker_agentes.csv"
    caminho_mapas = PASTA_SAIDA / "tracker_mapas.csv"
    caminho_excel = PASTA_SAIDA / "tracker_dados.xlsx"
    caminho_json = PASTA_SAIDA / "tracker_dados.json"

    df_agentes.to_csv(caminho_agentes, index=False, encoding="utf-8-sig")
    df_mapas.to_csv(caminho_mapas, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
        df_agentes.to_excel(writer, sheet_name="Agentes", index=False)
        df_mapas.to_excel(writer, sheet_name="Mapas", index=False)

        formatar_planilha(writer.sheets["Agentes"], df_agentes)
        formatar_planilha(writer.sheets["Mapas"], df_mapas)

    conteudo_json = {
        "jogador": RIOT_ID,
        "plataforma": PLATAFORMA,
        "playlist": PLAYLIST,
        "season_id": SEASON_ID,
        "coletado_em": datetime.now().astimezone().isoformat(timespec="seconds"),
        "urls": urls,
        "agentes": agentes,
        "mapas": mapas,
    }
    caminho_json.write_text(
        json.dumps(conteudo_json, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    print("\n" + "=" * 72)
    print("Arquivos criados:")
    for caminho in (
        caminho_agentes,
        caminho_mapas,
        caminho_excel,
        caminho_json,
    ):
        print(caminho.resolve())


def abrir_navegador(playwright, usar_chrome=False):
    argumentos = [
        "--disable-blink-features=AutomationControlled",
        "--start-maximized",
    ]

    if usar_chrome:
        return playwright.chromium.launch(
            channel="chrome",
            headless=MODO_HEADLESS,
            args=argumentos,
        )

    try:
        return playwright.chromium.launch(headless=MODO_HEADLESS, args=argumentos)
    except PlaywrightError as erro:
        print(f"Nao foi possivel abrir o Chromium: {erro}")
        print("Tentando usar o Google Chrome instalado...")
        return playwright.chromium.launch(
            channel="chrome",
            headless=MODO_HEADLESS,
            args=argumentos,
        )


def main():
    global RIOT_ID, SEASON_ID, PASTA_SAIDA

    parser = argparse.ArgumentParser(
        description="Coleta estatisticas de agentes e mapas do Tracker.gg."
    )
    parser.add_argument(
        "--usar-chrome",
        action="store_true",
        help="Usa o Google Chrome instalado em vez do Chromium do Playwright.",
    )
    parser.add_argument(
        "--riot-id",
        default=RIOT_ID,
        help="Riot ID no formato nome#tag.",
    )
    parser.add_argument(
        "--season-id",
        default=SEASON_ID,
        help="ID da season do Tracker.gg.",
    )
    parser.add_argument(
        "--saida",
        type=Path,
        default=PASTA_SAIDA,
        help="Pasta onde os arquivos da conta serao salvos.",
    )
    argumentos = parser.parse_args()

    RIOT_ID = argumentos.riot_id
    SEASON_ID = argumentos.season_id
    PASTA_SAIDA = argumentos.saida

    urls = montar_urls()
    print(f"Jogador: {RIOT_ID}")
    print(f"Season ID: {SEASON_ID}")

    with sync_playwright() as playwright:
        navegador = abrir_navegador(playwright, argumentos.usar_chrome)
        contexto = navegador.new_context(
            viewport={"width": 1_440, "height": 1_000},
            locale="pt-BR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/151.0.0.0 Safari/537.36"
            ),
        )
        page = contexto.new_page()

        try:
            agentes = extrair_pagina(page, urls["Agente"], "Agente")
            time.sleep(1)
            mapas = extrair_pagina(page, urls["Mapa"], "Mapa")
            salvar_resultados(agentes, mapas, urls)
        finally:
            contexto.close()
            navegador.close()

    if not agentes and not mapas:
        print("\nNenhuma tabela foi extraida. Consulte os arquivos debug_*.html.")
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
